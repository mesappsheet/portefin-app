from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Couleurs ---
BLUE_DARK  = "1F4E79"   # Titres principaux
BLUE_MID   = "2E75B6"   # Sous-groupes
BLUE_LIGHT = "BDD7EE"   # En-têtes colonnes
GREY_LIGHT = "F2F2F2"   # Lignes alternées
ORANGE     = "F4B942"   # Totaux
WHITE      = "FFFFFF"

thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, value, bold=True, bg=BLUE_MID, fg=WHITE, merge_cols=1, merge_rows=1,
        halign="center", valign="center", wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=9)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = border
    if merge_cols > 1 or merge_rows > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row + merge_rows - 1, end_column=col + merge_cols - 1
        )
    return cell

def data(ws, row, col, value, bold=False, bg=WHITE, halign="center", fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, size=9)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    cell.border = border
    if fmt:
        cell.number_format = fmt
    return cell

NUM = '#,##0'
PCT = '0.00'

# ==============================================================================
# ONGLET 1 — LES OUTPUTS
# ==============================================================================
ws = wb.active
ws.title = "LES OUTPUTS"
ws.sheet_view.showGridLines = False

# Titres
hdr(ws, 1, 1, "ACCES FINANCE BENIN SA", bg=WHITE, fg="000000", bold=True, merge_cols=22, halign="left")
hdr(ws, 2, 1, "PREVISIONS 2026", bg=WHITE, fg="000000", bold=True, merge_cols=22)
hdr(ws, 3, 1, "LES OUTPUTS", bg=WHITE, fg="555555", bold=True, merge_cols=22)

ROW_GRP  = 5
ROW_SUB  = 6
ROW_DATA = 7

# En-têtes de groupes
hdr(ws, ROW_GRP, 1,  "Mois",               bg=BLUE_DARK, merge_rows=2)
hdr(ws, ROW_GRP, 2,  "COMITES DE CREDIT",  bg=BLUE_DARK, merge_cols=5)
hdr(ws, ROW_GRP, 7,  "DEBOURSEMENTS",      bg=BLUE_DARK, merge_cols=3)
hdr(ws, ROW_GRP, 10, "RATIOS (%)",         bg=BLUE_DARK, merge_cols=3)
hdr(ws, ROW_GRP, 13, "AUTRES INDICATEURS", bg=BLUE_DARK, merge_cols=2)
hdr(ws, ROW_GRP, 15, "PRODUITS",           bg=BLUE_DARK, merge_cols=3)
hdr(ws, ROW_GRP, 18, "Remb atdu k",        bg=BLUE_DARK, merge_rows=2)
hdr(ws, ROW_GRP, 19, "Annuite atdu",       bg=BLUE_DARK, merge_rows=2)
hdr(ws, ROW_GRP, 20, "Ressource dispo",    bg=BLUE_DARK, merge_rows=2)
hdr(ws, ROW_GRP, 21, "montant impayé",     bg=BLUE_DARK, merge_rows=2)
hdr(ws, ROW_GRP, 22, "montant contentieux", bg=BLUE_DARK, merge_rows=2)

# Sous-en-têtes
sub_comites = ["dossier indivi", "dossier grpe", "TOTAL dts", "nbre comité", "montant accordé"]
for i, s in enumerate(sub_comites): hdr(ws, ROW_SUB, 2+i, s, bg=BLUE_LIGHT, fg="000000")

sub_deb = ["nbre clts débsés", "nbre déboursment", "montant déboursé"]
for i, s in enumerate(sub_deb): hdr(ws, ROW_SUB, 7+i, s, bg=BLUE_LIGHT, fg="000000")

sub_rat = ["Tx de Remb", "Tx de Res Ech", "Tx Imp PAR"]
for i, s in enumerate(sub_rat): hdr(ws, ROW_SUB, 10+i, s, bg=BLUE_LIGHT, fg="000000")

sub_aut = ["Nbre dts portefeuille", "Encours de crédit"]
for i, s in enumerate(sub_aut): hdr(ws, ROW_SUB, 13+i, s, bg=BLUE_LIGHT, fg="000000")

sub_prod = ["Revenus", "commission", "intérêt"]
for i, s in enumerate(sub_prod): hdr(ws, ROW_SUB, 15+i, s, bg=BLUE_LIGHT, fg="000000")

# Données (Exactement identiques à l'image)
# Mois, ind, grp, tot, com, mt_acc, cl_deb, nd_deb, mt_deb, remb, ech, par, portf, encours, rev, comm, int, r_atdu, a_atdu, res_dispo, impaye, cont
data_rows = [
    ("Janv", 7, 9, 16, 1, 82516800, 16, 2, 81591632, 98, 98, 4, 330, 948691297, 18819112, 816916, 18002196, 155170033, 173172229, 98478401, 3103401, 5876525),
    ("Févr", 33, 24, 57, 4, 206818200, 56, 4, 203750018, 98, 98, 4, 368, 1086593928, 19672662, 3037500, 16635162, 162626455, 179281617, 57354839, 3252529, 5758994),
    ("Mars", 49, 24, 73, 4, 404354200, 72, 4, 400310658, 98, 98, 4, 415, 1313796857, 23179062, 4003107, 19175945, 169740020, 188915965, -173215799, 3394800, 5643814),
    ("Avr", 43, 35, 78, 4, 336567600, 78, 4, 333201924, 98, 98, 4, 471, 1469956305, 26676807, 3332019, 23344788, 173534800, 196879587, -332882923, 3470696, 5530938),
    ("Mai", 35, 35, 70, 4, 300644200, 70, 4, 297637758, 98, 98, 4, 503, 1582742898, 29197731, 2976378, 26221354, 181269850, 207491204, -449250831, 3625397, 5420319),
    ("Juin", 35, 44, 79, 4, 209567900, 78, 4, 207472221, 98, 98, 4, 555, 1600548377, 30382480, 2074722, 28307758, 185932938, 214240696, -470790114, 3718659, 5311913),
    ("Juil", 36, 29, 65, 4, 201781000, 65, 4, 198763190, 98, 98, 4, 591, 1700500119, 31363983, 2987632, 28376351, 194986551, 223362902, -472889256, 3899731, 5205674),
    ("Août", 29, 36, 65, 4, 249821400, 64, 4, 247323186, 98, 98, 4, 639, 1742083780, 32649128, 2473232, 30175897, 201735681, 231911577, -520154259, 4034714, 5101561),
    ("Sept", 30, 56, 86, 4, 195509800, 85, 4, 193554702, 98, 98, 4, 680, 1717961219, 32859239, 1935547, 30923692, 213540518, 244464210, -500168442, 4270810, 4999530),
    ("Oct", 34, 47, 81, 4, 197827500, 80, 4, 195849225, 98, 98, 4, 730, 1686311231, 32443217, 1958492, 30484725, 223128412, 253613136, -473889256, 4462568, 4899539),
    ("Nov", 31, 27, 58, 4, 187032800, 57, 4, 185162472, 98, 98, 4, 752, 1635791980, 31762340, 1851625, 29910715, 231121165, 261031880, -176930563, 4622423, 4801548),
    ("Déc", 14, 15, 29, 2, 136652800, 28, 4, 135286272, 98, 98, 4, 752, 1534324853, 30349798, 1352863, 28996935, 232034944, 261031880, -80181891, 4640699, 4705517),
]

for r_idx, r in enumerate(data_rows):
    cur_row = ROW_DATA + r_idx
    bg = GREY_LIGHT if r_idx % 2 == 0 else WHITE
    data(ws, cur_row, 1, r[0], bold=True, bg=bg, halign="left")
    for c_idx, val in enumerate(r[1:]):
        fmt = NUM if c_idx not in [8,9,10] else PCT
        data(ws, cur_row, 2+c_idx, val, bg=bg, fmt=fmt)

# Total Outputs
tot_row = ROW_DATA + len(data_rows)
hdr(ws, tot_row, 1, "Total", bg=ORANGE, merge_rows=1)
tot_vals = [372, 385, 757, 43, 2909094200, 749, 46, 2880003258, None, None, None, None, None, 339355549, 28800033, 310555516, 2324821367, 2635376884, None, None, None]
for i, v in enumerate(tot_vals):
    if v is not None:
        c = data(ws, tot_row, 2+i, v, bold=True, bg=ORANGE, fmt=NUM)

# Note bf=
bf_row = tot_row + 1
ws.cell(row=bf_row, column=18, value="bf=")
ws.cell(row=bf_row, column=19, value=530181891).number_format = NUM

# ==============================================================================
# ONGLET 2 — LES INPUTS
# ==============================================================================
ws2 = wb.create_sheet("LES INPUTS")
ws2.sheet_view.showGridLines = False

hdr(ws2, 1, 1, "LES INPUTS", bg=WHITE, fg="000000", bold=True, merge_cols=23, halign="left")

hdrs_inputs = ["ind /cp/ct", "é", "cp", "nb cté", "a rea n-1", "ta", "n group /cp", "ga réal/n-1", "mia réa/n-1", "tmai", "mng", "mga", "tga", "min", "D", "td", "Nouv CLIT IND", "Nouv CLIT grou", "mt accdé IND/N", "mt accdé grp/N", "mt accdé IND/A", "mt accdé grp/A", "FG/N"]
for i, h in enumerate(hdrs_inputs): hdr(ws2, 3, i+1, h, bg=BLUE_DARK)

input_rows = [
    (1, 2, 1, 5, 0.98, 1, 1, 61000000, 1.3, 75000, 300000, 1.2, 2000000, 2, 0.99, 2, 2, 4000000, 450000, 77714000, 352800, 440550, None),
    (2, 2, 4, 17, 0.98, 1, 0, 214300000, 1.3, 75000, 0, 1.2, 2000000, 4, 0.99, 16, 6, 32000000, 1350000, 273018200, 0, 3301650, None),
    (3, 2, 4, 25, 0.98, 1, 0, 278300000, 1.3, 75000, 0, 1.2, 2000000, 4, 0.99, 24, 8, 48000000, 1800000, 354554200, 0, 4930200, None),
    (3, 2, 4, 19, 0.98, 1, 4, 223500000, 1.3, 75000, 1725000, 1.2, 2000000, 4, 0.99, 24, 8, 48000000, 1800000, 284739000, 2028600, 4930200, None),
    (2, 2, 4, 19, 0.98, 1, 4, 208000000, 1.3, 75000, 1575000, 1.2, 2000000, 4, 0.99, 16, 8, 32000000, 1800000, 264430000, 1852200, 3346200, None),
    (2, 2, 4, 19, 0.98, 1, 7, 135750000, 1.3, 75000, 2400000, 1.2, 2000000, 4, 0.99, 16, 8, 32000000, 1800000, 172945500, 2822400, 3346200, None),
    (2, 2, 4, 20, 0.98, 1, 2, 210000000, 1.3, 75000, 375000, 1.2, 2000000, 4, 0.99, 16, 8, 32000000, 1800000, 267540000, 441000, 3346200, None),
    (2, 2, 4, 13, 0.98, 1, 4, 168800000, 1.3, 75000, 825000, 1.2, 2000000, 4, 0.99, 16, 8, 32000000, 1800000, 215051200, 970200, 3346200, None),
    (2, 2, 4, 14, 0.98, 1, 11, 124300000, 1.3, 75000, 2850000, 1.2, 2000000, 4, 0.99, 16, 8, 32000000, 1800000, 158358200, 3351600, 3346200, None),
    (2, 2, 4, 18, 0.98, 1, 8, 126950000, 1.3, 75000, 1950000, 1.2, 2000000, 4, 0.99, 16, 8, 32000000, 1800000, 161734300, 2293200, 3346200, None),
    (2, 2, 4, 15, 0.98, 1, 1, 120000000, 1.3, 75000, 300000, 1.2, 2000000, 4, 0.99, 16, 8, 32000000, 1800000, 152880000, 352800, 3346200, None),
    (1, 2, 2, 10, 0.98, 1, 1, 100000000, 1.3, 75000, 300000, 1.2, 2000000, 4, 0.99, 4, 4, 8000000, 900000, 127400000, 352800, 881100, None),
]

for r_idx, r in enumerate(input_rows):
    cur_row = 4 + r_idx
    bg = GREY_LIGHT if r_idx % 2 == 0 else WHITE
    for c_idx, val in enumerate(r):
        fmt = NUM if (val and isinstance(val, (int, float)) and val > 100) else None
        if c_idx == 4: fmt = "0%"
        data(ws2, cur_row, c_idx+1, val, bg=bg, fmt=fmt)

# Total Inputs
tot_row_in = 4 + len(input_rows)
hdr(ws2, tot_row_in, 1, "Total", bg=ORANGE, merge_cols=16)
tot_vals_in = [182, 84, 364000000, 18900000, 2510926600, 14817600, 37907100]
for i, v in enumerate(tot_vals_in):
    data(ws2, tot_row_in, 17+i, v, bold=True, bg=ORANGE, fmt=NUM)

# Formatting
for w in range(1, 24): ws2.column_dimensions[get_column_letter(w)].width = 12
for w in range(1, 23): ws.column_dimensions[get_column_letter(w)].width = 13
ws.column_dimensions['A'].width = 8

wb.save("PREVISIONS_2026.xlsx")
print("Fichier PREVISIONS_2026.xlsx mis à jour.")
